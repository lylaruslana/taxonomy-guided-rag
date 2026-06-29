"""
grid_search_ce_weights.py — IC3INA 2026
Grid search bobot reranking (vector, cross encoder, taxonomy).
Cross encoder (bge-reranker-v2-m3) menggantikan BM25 sebagai sinyal reranking.

Alur:
  1. Load 71 pertanyaan + reference answer dari qa-dokter-eg.xlsx
  2. Load taxonomy
  3. Embed queries dengan BGE-M3 (GPU)
  4. Query ChromaDB → top-K kandidat
  5. Load BAAI/bge-reranker-v2-m3 → hitung CE score semua (query, chunk) pairs
     → Cache ke ce_scores_cache.json; skip inference jika cache valid sudah ada
  6. Hitung taxonomy score (via taxonomy_runtime)
  7. Grid search: semua kombinasi (w_vector, w_cross, w_taxonomy), step 0.1
     final = w_v*vector + w_c*cross_encoder + w_t*taxonomy
  8. Visualisasi: histogram distribusi P@5 + heatmap (w_cross vs w_taxonomy) → PNG
  9. Laporan top-20 + simpan CSV

Jalankan dari root project:
    python3 -u grid_search_ce_weights.py

Output:
    grid_search_ce_results.csv
    grid_search_ce_results.png
    ce_scores_cache.json   (cache CE, skip model load jika sudah ada)
"""

import json, sys, os, re, csv, time
from pathlib import Path
from itertools import product

ROOT = Path(__file__).parent

# ---------------------------------------------------------------------------
# KONFIGURASI
# ---------------------------------------------------------------------------
CHROMA_DB    = str(ROOT / "chromaDB/D003D-TAX-QEXP-CLEAN-bgem3-sahabatai-semantic")
COLLECTION   = "D003D-TAX-QEXP-CLEAN-bgem3-sahabatai-semantic"
BGE_MODEL    = str(ROOT / "local_bge_m3")
CE_MODEL     = "BAAI/bge-reranker-v2-m3"   # ganti ke path lokal jika sudah didownload
TEST_FILE    = str(ROOT / "data/data_test/qa-dokter-eg.xlsx")
TAX_CONCEPTS = str(ROOT / "data/taxonomy_concepts_v4.json")
SYN_INDEX    = str(ROOT / "data/synonym_index_v4.json")
OUTPUT_CSV   = str(ROOT / "grid_search_ce_results.csv")
OUTPUT_PNG   = str(ROOT / "grid_search_ce_results.png")
CE_CACHE     = str(ROOT / "ce_scores_cache.json")
HF_TOKEN = os.environ.get("HF_TOKEN", "")

RETRIEVAL_K  = 20    # kandidat dari ChromaDB
TOP_K        = 5     # chunk final dievaluasi
WEIGHT_STEP  = 0.1
CE_BATCH     = 32    # (query, chunk) pairs per batch

# Referensi perbandingan: hasil grid search BM25 terbaik (dari grid_search_results.csv)
BM25_BEST_P5 = 0.8986   # w_v=0.50, w_b=0.40, w_t=0.10

# ---------------------------------------------------------------------------
# [1/7] LOAD DATA
# ---------------------------------------------------------------------------
print("[1/7] Load test questions...", flush=True)
import openpyxl
wb = openpyxl.load_workbook(TEST_FILE)
ws = wb.active
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
col_q   = headers.index("Pertanyaan") + 1
col_ans = headers.index("Jawaban") + 1
questions, references = [], []
for r in range(2, ws.max_row + 1):
    q = ws.cell(r, col_q).value
    a = ws.cell(r, col_ans).value
    if q and a:
        questions.append(str(q).strip())
        references.append(str(a).strip())
print(f"  {len(questions)} pertanyaan dimuat.", flush=True)

# ---------------------------------------------------------------------------
# [2/7] LOAD TAXONOMY
# ---------------------------------------------------------------------------
print("[2/7] Load taxonomy...", flush=True)
from taxonomy_runtime import (
    load_taxonomy_concepts, load_synonym_index,
    detect_query_concepts, compute_taxonomy_score,
)
taxonomy_by_id = load_taxonomy_concepts(TAX_CONCEPTS)
synonym_index  = load_synonym_index(SYN_INDEX)
print(f"  {len(taxonomy_by_id)} konsep, {len(synonym_index)} sinonim.", flush=True)

# ---------------------------------------------------------------------------
# [3/7] EMBED QUERIES DENGAN BGE-M3 (GPU)
# ---------------------------------------------------------------------------
print("[3/7] Embed queries dengan BGE-M3 (GPU)...", flush=True)
import torch
import numpy as np
from transformers import AutoTokenizer, AutoModel

device = "cuda" if torch.cuda.is_available() else "cpu"
print(f"  Device: {device}" + (f" ({torch.cuda.get_device_name(0)})" if device == "cuda" else ""), flush=True)

os.environ["HUGGING_FACE_HUB_TOKEN"] = HF_TOKEN
os.environ["HF_TOKEN"] = HF_TOKEN

_bge_tok = AutoTokenizer.from_pretrained(BGE_MODEL)
_bge_mod = AutoModel.from_pretrained(BGE_MODEL).to(device)
_bge_mod.eval()

def _embed_batch(texts, batch_size=32):
    all_vecs = []
    for i in range(0, len(texts), batch_size):
        batch = texts[i:i + batch_size]
        enc   = _bge_tok(batch, padding=True, truncation=True, max_length=512, return_tensors="pt")
        enc   = {k: v.to(device) for k, v in enc.items()}
        with torch.no_grad():
            out = _bge_mod(**enc)
        mask = enc["attention_mask"].unsqueeze(-1).float()
        vecs = (out.last_hidden_state * mask).sum(1) / mask.sum(1)
        vecs = torch.nn.functional.normalize(vecs, p=2, dim=1)
        all_vecs.append(vecs.cpu().numpy())
        print(f"  embed {min(i+batch_size, len(texts))}/{len(texts)}", flush=True)
    return np.vstack(all_vecs)

query_embeddings = _embed_batch(questions)
print(f"  Done. shape={query_embeddings.shape}", flush=True)

# Bebaskan GPU setelah embedding
del _bge_mod
import gc
gc.collect()
if device == "cuda":
    torch.cuda.empty_cache()
    print(f"  VRAM setelah free BGE-M3: {torch.cuda.memory_allocated()/1e9:.2f} GB", flush=True)

# ---------------------------------------------------------------------------
# [4/7] QUERY CHROMADB
# ---------------------------------------------------------------------------
print(f"[4/7] Query ChromaDB (top-{RETRIEVAL_K} per query)...", flush=True)
import chromadb
chroma_client = chromadb.PersistentClient(path=CHROMA_DB)
col           = chroma_client.get_collection(COLLECTION)

all_candidates = []
for i, (q, emb) in enumerate(zip(questions, query_embeddings)):
    res = col.query(
        query_embeddings=[emb.tolist()],
        n_results=RETRIEVAL_K,
        include=["documents", "metadatas", "distances"],
    )
    candidates = []
    for doc, meta, dist in zip(res["documents"][0], res["metadatas"][0], res["distances"][0]):
        vec_score = max(0.0, 1.0 - float(dist))
        tags = []
        raw = meta.get("taxonomy_tags")
        if raw:
            try:
                tags = json.loads(raw) if isinstance(raw, str) else raw
            except Exception:
                tags = []
        candidates.append({
            "text": doc,
            "taxonomy_tags": tags,
            "vector_score": vec_score,
        })
    all_candidates.append(candidates)
    if (i + 1) % 10 == 0:
        print(f"  {i+1}/{len(questions)}", flush=True)
print("  ChromaDB query selesai.", flush=True)

# ---------------------------------------------------------------------------
# [5/7] CROSS ENCODER SCORES — load atau dari cache
# ---------------------------------------------------------------------------
print("[5/7] Cross encoder scores (bge-reranker-v2-m3)...", flush=True)
from transformers import AutoModelForSequenceClassification

# Cek cache: valid jika model dan collection cocok
_cache_hit = False
_cache_data = {}
if Path(CE_CACHE).exists():
    try:
        with open(CE_CACHE, encoding="utf-8") as _f:
            _cache_data = json.load(_f)
        if (
            _cache_data.get("model") == CE_MODEL
            and _cache_data.get("collection") == COLLECTION
            and len(_cache_data.get("scores", [])) == len(questions)
        ):
            _cache_hit = True
            print(f"  Cache valid ditemukan → skip inference. ({CE_CACHE})", flush=True)
        else:
            print(f"  Cache ada tapi tidak cocok → re-inference.", flush=True)
    except Exception as e:
        print(f"  Cache rusak ({e}) → re-inference.", flush=True)

if _cache_hit:
    # scores[query_idx][cand_idx] = sigmoid(logit)
    ce_scores_all = _cache_data["scores"]
else:
    print(f"  Load model {CE_MODEL} ke {device}...", flush=True)
    _ce_tok = AutoTokenizer.from_pretrained(CE_MODEL, token=HF_TOKEN)
    _ce_mod = AutoModelForSequenceClassification.from_pretrained(
        CE_MODEL, token=HF_TOKEN
    ).to(device)
    _ce_mod.eval()
    if device == "cuda":
        print(f"  VRAM setelah load CE: {torch.cuda.memory_allocated()/1e9:.2f} GB", flush=True)

    t_ce_start = time.time()
    ce_scores_all = []

    for i, (q, cands) in enumerate(zip(questions, all_candidates)):
        pairs     = [[q, c["text"]] for c in cands]
        q_scores  = []
        for b_start in range(0, len(pairs), CE_BATCH):
            batch_pairs = pairs[b_start:b_start + CE_BATCH]
            enc = _ce_tok(
                batch_pairs,
                padding=True,
                truncation=True,
                max_length=512,
                return_tensors="pt",
            )
            enc = {k: v.to(device) for k, v in enc.items()}
            with torch.no_grad():
                logits = _ce_mod(**enc, return_dict=True).logits.view(-1).float()
                # sigmoid normalisasi ke [0, 1]
                scores = torch.sigmoid(logits).cpu().tolist()
            q_scores.extend(scores)
        ce_scores_all.append(q_scores)
        if (i + 1) % 10 == 0:
            elapsed = round(time.time() - t_ce_start, 1)
            print(f"  CE inference {i+1}/{len(questions)} ({elapsed}s)", flush=True)

    elapsed_total = round(time.time() - t_ce_start, 1)
    print(f"  CE inference selesai: {elapsed_total}s untuk {len(questions)} query × {RETRIEVAL_K} kandidat", flush=True)

    # Simpan cache
    _cache_data = {
        "model": CE_MODEL,
        "collection": COLLECTION,
        "n_queries": len(questions),
        "n_candidates": RETRIEVAL_K,
        "scores": ce_scores_all,
    }
    with open(CE_CACHE, "w", encoding="utf-8") as _f:
        json.dump(_cache_data, _f, ensure_ascii=False)
    print(f"  Cache disimpan → {CE_CACHE}", flush=True)

    # Bebaskan GPU
    del _ce_mod
    gc.collect()
    if device == "cuda":
        torch.cuda.empty_cache()
        print(f"  VRAM setelah free CE model: {torch.cuda.memory_allocated()/1e9:.2f} GB", flush=True)

# Attach CE scores ke candidates
for i, cands in enumerate(all_candidates):
    for j, cand in enumerate(cands):
        cand["ce_score"] = float(ce_scores_all[i][j]) if j < len(ce_scores_all[i]) else 0.0

# ---------------------------------------------------------------------------
# [6/7] TAXONOMY SCORE + PROXY LABEL
# ---------------------------------------------------------------------------
print("[6/7] Hitung taxonomy score + proxy label...", flush=True)

def tokenize(text: str) -> list:
    return re.findall(r"\b\w+\b", text.lower())

def proxy_relevance(chunk_text: str, reference: str) -> float:
    """F1 token overlap chunk vs reference answer (proxy relevance label)."""
    chunk_tok = set(tokenize(chunk_text))
    ref_tok   = set(tokenize(reference))
    if not chunk_tok or not ref_tok:
        return 0.0
    common = chunk_tok & ref_tok
    if not common:
        return 0.0
    precision = len(common) / len(chunk_tok)
    recall    = len(common) / len(ref_tok)
    return 2 * precision * recall / (precision + recall)

for i, (q, ref, cands) in enumerate(zip(questions, references, all_candidates)):
    q_concepts = detect_query_concepts(q, synonym_index, taxonomy_by_id)
    for cand in cands:
        cand["taxonomy_score"] = float(compute_taxonomy_score(cand["taxonomy_tags"], q_concepts))
        cand["proxy"]          = proxy_relevance(cand["text"], ref)

print("  Done.", flush=True)

# ---------------------------------------------------------------------------
# [7/7] GRID SEARCH
# ---------------------------------------------------------------------------
print("[7/7] Grid search bobot (vector, cross_encoder, taxonomy)...", flush=True)

# 0.0 diikutkan agar triangel penuh → 66 kombinasi (vs 45 di BM25 version)
weight_values = [round(v * WEIGHT_STEP, 2) for v in range(0, int(1.0 / WEIGHT_STEP) + 1)]

combos = []
for wv, wc in product(weight_values, repeat=2):
    wt = round(1.0 - wv - wc, 2)
    if wt < -0.001 or wt > 1.001:
        continue
    wt = max(0.0, wt)
    if round(wv + wc + wt, 2) < 0.99:
        continue
    combos.append((wv, wc, round(wt, 2)))

print(f"  Total kombinasi: {len(combos)}", flush=True)

PROXY_THRESHOLD = 0.1

def evaluate_weights(w_v, w_c, w_t):
    p5_list = []
    for cands, ref in zip(all_candidates, references):
        scored = []
        for c in cands:
            fs = w_v * c["vector_score"] + w_c * c["ce_score"] + w_t * c["taxonomy_score"]
            scored.append((fs, c["proxy"]))
        scored.sort(key=lambda x: x[0], reverse=True)
        top5 = scored[:TOP_K]
        p5   = sum(1 for _, prx in top5 if prx >= PROXY_THRESHOLD) / TOP_K
        p5_list.append(p5)
    return sum(p5_list) / len(p5_list)

results = []
for idx, (wv, wc, wt) in enumerate(combos):
    score = evaluate_weights(wv, wc, wt)
    results.append((score, wv, wc, wt))
    if (idx + 1) % 20 == 0:
        print(f"  {idx+1}/{len(combos)} kombinasi...", flush=True)

results.sort(reverse=True)
best_score, best_wv, best_wc, best_wt = results[0]
print(f"\n  Terbaik: w_v={best_wv:.2f}  w_c={best_wc:.2f}  w_t={best_wt:.2f}  P@5={best_score:.4f}", flush=True)
print(f"  BM25 best (referensi): P@5={BM25_BEST_P5:.4f}", flush=True)
delta = best_score - BM25_BEST_P5
print(f"  Delta CE vs BM25: {delta:+.4f} ({'lebih baik' if delta > 0 else 'lebih rendah'})", flush=True)

# ---------------------------------------------------------------------------
# VISUALISASI — histogram distribusi P@5 + heatmap w_cross vs w_taxonomy
# ---------------------------------------------------------------------------
try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.colors import Normalize
    import matplotlib.patches as mpatches
    HAS_MPL = True
except ImportError:
    HAS_MPL = False
    print("\n[WARN] matplotlib tidak tersedia → skip visualisasi.", flush=True)

if HAS_MPL:
    print("\nGenerasi plot...", flush=True)
    all_scores = [s for s, *_ in results]
    score_map  = {(wc, wt): s for s, wv, wc, wt in results}

    # --- Heatmap array ---
    wc_vals = sorted({wc for _, _, wc, _ in results})
    wt_vals = sorted({wt for _, _, _, wt in results})
    hm = np.full((len(wt_vals), len(wc_vals)), np.nan)
    best_i, best_j = 0, 0
    for i, wt in enumerate(wt_vals):
        for j, wc in enumerate(wc_vals):
            key = (round(wc, 2), round(wt, 2))
            if key in score_map:
                hm[i, j] = score_map[key]
                if abs(wc - best_wc) < 0.001 and abs(wt - best_wt) < 0.001:
                    best_i, best_j = i, j

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13, 5))
    fig.suptitle(
        f"Grid Search Cross Encoder — bge-reranker-v2-m3\n"
        f"Best: w_vector={best_wv:.2f}  w_cross={best_wc:.2f}  w_taxonomy={best_wt:.2f}  "
        f"P@5={best_score:.4f}  (BM25 best: {BM25_BEST_P5:.4f})",
        fontsize=11, y=1.02
    )

    # --- ax1: Histogram ---
    ax1.hist(all_scores, bins=15, color="#4C72B0", edgecolor="white", alpha=0.85)
    ax1.axvline(BM25_BEST_P5, color="#DD8452", linestyle="--", linewidth=1.5,
                label=f"BM25 best ({BM25_BEST_P5:.4f})")
    ax1.axvline(best_score,   color="#55A868", linestyle="--", linewidth=1.5,
                label=f"CE best ({best_score:.4f})")
    ax1.set_xlabel("Proxy P@5", fontsize=10)
    ax1.set_ylabel("Jumlah kombinasi bobot", fontsize=10)
    ax1.set_title("Distribusi P@5 semua kombinasi bobot", fontsize=10)
    ax1.legend(fontsize=9)
    ax1.grid(axis="y", alpha=0.3)

    # --- ax2: Heatmap ---
    vmin = float(np.nanmin(hm))
    vmax = float(np.nanmax(hm))
    norm = Normalize(vmin=vmin, vmax=vmax)

    # Grey background untuk sel invalid (NaN)
    ax2.set_facecolor("#DDDDDD")
    im = ax2.imshow(
        hm, origin="lower", aspect="auto",
        extent=[-0.05, len(wc_vals) - 0.95, -0.05, len(wt_vals) - 0.95],
        norm=norm, cmap="YlOrRd", interpolation="nearest",
    )
    # Anotasi nilai di setiap sel valid
    for i in range(len(wt_vals)):
        for j in range(len(wc_vals)):
            if not np.isnan(hm[i, j]):
                txt_color = "black" if hm[i, j] < (vmin + (vmax - vmin) * 0.7) else "white"
                ax2.text(j, i, f"{hm[i,j]:.3f}", ha="center", va="center",
                         fontsize=6.5, color=txt_color)

    # Tandai sel terbaik dengan border merah
    ax2.add_patch(plt.Rectangle(
        (best_j - 0.5, best_i - 0.5), 1, 1,
        fill=False, edgecolor="red", linewidth=2.0, zorder=5,
    ))

    cbar = fig.colorbar(im, ax=ax2, shrink=0.85)
    cbar.set_label("Proxy P@5", fontsize=9)

    ax2.set_xticks(range(len(wc_vals)))
    ax2.set_xticklabels([f"{v:.1f}" for v in wc_vals], fontsize=8)
    ax2.set_yticks(range(len(wt_vals)))
    ax2.set_yticklabels([f"{v:.1f}" for v in wt_vals], fontsize=8)
    ax2.set_xlabel("w_cross", fontsize=10)
    ax2.set_ylabel("w_taxonomy", fontsize=10)
    ax2.set_title("Proxy P@5 (w_vector = 1 − w_cross − w_taxonomy)\nArea abu = invalid (w_vector < 0)", fontsize=9)

    plt.tight_layout()
    plt.savefig(OUTPUT_PNG, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  Plot disimpan → {OUTPUT_PNG}", flush=True)

# ---------------------------------------------------------------------------
# LAPORAN
# ---------------------------------------------------------------------------
print("\n" + "=" * 70)
print(f"{'Rank':<5} {'w_vector':>9} {'w_cross':>8} {'w_tax':>6} {'P@5':>8}")
print("-" * 70)

current_rank = next(
    (r for r, (s, wv, wc, wt) in enumerate(results, 1) if abs(wv - best_wv) < 0.001
     and abs(wc - best_wc) < 0.001 and abs(wt - best_wt) < 0.001),
    None,
)

print(f"[CE BEST]  w_v={best_wv:.2f}  w_c={best_wc:.2f}  w_t={best_wt:.2f}  P@5={best_score:.4f}")
print(f"[BM25 REF] w_v=0.50  w_b=0.40  w_t=0.10  P@5={BM25_BEST_P5:.4f}  (grid_search_results.csv)")
print(f"[DELTA]    {delta:+.4f}")
print("-" * 70)
for rank, (score, wv, wc, wt) in enumerate(results[:20], 1):
    marker = " ← BEST" if rank == 1 else ""
    print(f"{rank:<5} {wv:>9.2f} {wc:>8.2f} {wt:>6.2f} {score:>8.4f}{marker}")

# ---------------------------------------------------------------------------
# SIMPAN CSV
# ---------------------------------------------------------------------------
with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(["rank", "w_vector", "w_cross", "w_taxonomy", "proxy_p5"])
    for rank, (score, wv, wc, wt) in enumerate(results, 1):
        writer.writerow([rank, wv, wc, wt, f"{score:.4f}"])

print(f"\n[SAVED] {OUTPUT_CSV}")
print("[DONE]")
