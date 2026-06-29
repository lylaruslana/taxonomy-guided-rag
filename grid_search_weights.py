"""
grid_search_weights.py — IC3INA 2026
Grid search bobot reranking (vector, BM25, taxonomy) secara offline.

Alur:
  1. Load 71 pertanyaan + reference answer dari qa-dokter-eg.xlsx
  2. Embed tiap pertanyaan dengan BGE-M3 (local)
  3. Query ChromaDB → top-K kandidat beserta teks & metadata
  4. Hitung BM25 score (lokal, atas K kandidat)
  5. Hitung taxonomy score (via taxonomy_runtime)
  6. Grid search: coba semua kombinasi bobot
     final = w_v*vector + w_b*bm25 + w_t*taxonomy
  7. Evaluasi proxy: token overlap top-5 vs reference answer → Precision@5
  8. Laporan top-20 kombinasi terbaik

Jalankan dari root project:
    python grid_search_weights.py

Output:
    grid_search_results.csv
"""

import json, sys, os, math, re, csv
from pathlib import Path
from itertools import product

ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT))

# ---------------------------------------------------------------------------
# KONFIGURASI
# ---------------------------------------------------------------------------
CHROMA_DB    = str(ROOT / "chromaDB/D003D-TAX-QEXP-CLEAN-bgem3-sahabatai-semantic")
COLLECTION   = "D003D-TAX-QEXP-CLEAN-bgem3-sahabatai-semantic"
BGE_MODEL    = str(ROOT / "local_bge_m3")
TEST_FILE    = str(ROOT / "data/data_test/qa-dokter-eg.xlsx")
TAX_CONCEPTS = str(ROOT / "data/taxonomy_concepts_v4.json")
SYN_INDEX    = str(ROOT / "data/synonym_index_v4.json")
OUTPUT_CSV   = str(ROOT / "grid_search_results.csv")

RETRIEVAL_K  = 20    # kandidat yang diambil dari ChromaDB
TOP_K        = 5     # chunk final yang dievaluasi
WEIGHT_STEP  = 0.1   # granularitas grid (0.1 = 11 nilai per dimensi)

# Bobot saat ini (baseline perbandingan)
CURRENT_W = (0.45, 0.25, 0.30)  # vector, bm25, taxonomy

# ---------------------------------------------------------------------------
# LOAD DATA
# ---------------------------------------------------------------------------
print("[1/6] Load test questions...", flush=True)
import openpyxl
wb = openpyxl.load_workbook(TEST_FILE)
ws = wb.active
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
col_q   = headers.index("Pertanyaan") + 1
col_ans = headers.index("Jawaban") + 1
questions, references = [], []
for r in range(2, ws.max_row + 1):
    q = ws.cell(r, col_q).value
    a = ws.cell(r, col_ans).value
    if q and a:
        questions.append(str(q).strip())
        references.append(str(a).strip())
print(f"  {len(questions)} pertanyaan dimuat.", flush=True)

# ---------------------------------------------------------------------------
# LOAD TAXONOMY
# ---------------------------------------------------------------------------
print("[2/6] Load taxonomy...", flush=True)
from taxonomy_runtime import (
    load_taxonomy_concepts, load_synonym_index,
    detect_query_concepts, compute_taxonomy_score,
)
taxonomy_by_id = load_taxonomy_concepts(TAX_CONCEPTS)
synonym_index  = load_synonym_index(SYN_INDEX)
print(f"  {len(taxonomy_by_id)} konsep, {len(synonym_index)} sinonim.", flush=True)

# ---------------------------------------------------------------------------
# LOAD BGE-M3 + EMBED QUERIES
# ---------------------------------------------------------------------------
print("[3/6] Embed queries dengan BGE-M3...", flush=True)
import torch
from transformers import AutoTokenizer, AutoModel
import numpy as np

_tokenizer = AutoTokenizer.from_pretrained(BGE_MODEL)
_model     = AutoModel.from_pretrained(BGE_MODEL)
_model.eval()

def _embed_batch(texts, batch_size=16):
    all_vecs = []
    for i in range(0, len(texts), batch_size):
        batch = texts[i:i + batch_size]
        enc   = _tokenizer(batch, padding=True, truncation=True, max_length=512, return_tensors="pt")
        with torch.no_grad():
            out = _model(**enc)
        # mean pooling
        mask = enc["attention_mask"].unsqueeze(-1).float()
        vecs = (out.last_hidden_state * mask).sum(1) / mask.sum(1)
        # L2 normalize
        vecs = torch.nn.functional.normalize(vecs, p=2, dim=1)
        all_vecs.append(vecs.cpu().numpy())
        print(f"  embed {min(i+batch_size, len(texts))}/{len(texts)}", flush=True)
    return np.vstack(all_vecs)

query_embeddings = _embed_batch(questions)
print(f"  Done. shape={query_embeddings.shape}", flush=True)

# ---------------------------------------------------------------------------
# QUERY CHROMADB — ambil top-K per query
# ---------------------------------------------------------------------------
print(f"[4/6] Query ChromaDB (top-{RETRIEVAL_K} per query)...", flush=True)
import chromadb
client = chromadb.PersistentClient(path=CHROMA_DB)
col    = client.get_collection(COLLECTION)

all_candidates = []   # list of list: per query, list of candidate dicts
for i, (q, emb) in enumerate(zip(questions, query_embeddings)):
    res = col.query(
        query_embeddings=[emb.tolist()],
        n_results=RETRIEVAL_K,
        include=["documents", "metadatas", "distances"],
    )
    candidates = []
    docs      = res["documents"][0]
    metas     = res["metadatas"][0]
    distances = res["distances"][0]
    for doc, meta, dist in zip(docs, metas, distances):
        # ChromaDB cosine distance → similarity
        vec_score = max(0.0, 1.0 - dist)
        tags = []
        raw = meta.get("taxonomy_tags")
        if raw:
            try:
                tags = json.loads(raw) if isinstance(raw, str) else raw
            except Exception:
                tags = []
        candidates.append({
            "text": doc,
            "taxonomy_tags": tags,
            "vector_score": vec_score,
        })
    all_candidates.append(candidates)
    if (i + 1) % 10 == 0:
        print(f"  {i+1}/{len(questions)}", flush=True)
print("  ChromaDB query selesai.", flush=True)

# ---------------------------------------------------------------------------
# HITUNG BM25 + TAXONOMY SCORE per query
# ---------------------------------------------------------------------------
print("[5/6] Hitung BM25 & taxonomy scores...", flush=True)
from rank_bm25 import BM25Okapi

def tokenize(text: str) -> list:
    return re.findall(r"\b\w+\b", text.lower())

for i, (q, cands) in enumerate(zip(questions, all_candidates)):
    # BM25 atas K kandidat
    corpus = [tokenize(c["text"]) for c in cands]
    bm25   = BM25Okapi(corpus)
    q_tok  = tokenize(q)
    raw_scores = bm25.get_scores(q_tok)
    max_bm25   = max(raw_scores) if max(raw_scores) > 0 else 1.0
    norm_bm25  = [s / max_bm25 for s in raw_scores]

    # Taxonomy
    q_concepts = detect_query_concepts(q, synonym_index, taxonomy_by_id)
    for j, cand in enumerate(cands):
        cand["bm25_score"]     = float(norm_bm25[j])
        cand["taxonomy_score"] = float(compute_taxonomy_score(cand["taxonomy_tags"], q_concepts))

print("  Done.", flush=True)

# ---------------------------------------------------------------------------
# PROXY LABEL: token overlap tiap chunk vs reference answer
# ---------------------------------------------------------------------------
def proxy_relevance(chunk_text: str, reference: str) -> float:
    """F1 token overlap antara chunk dan reference answer."""
    chunk_tok = set(tokenize(chunk_text))
    ref_tok   = set(tokenize(reference))
    if not chunk_tok or not ref_tok:
        return 0.0
    common    = chunk_tok & ref_tok
    if not common:
        return 0.0
    precision = len(common) / len(chunk_tok)
    recall    = len(common) / len(ref_tok)
    return 2 * precision * recall / (precision + recall)

# Precompute proxy scores
for i, (ref, cands) in enumerate(zip(references, all_candidates)):
    for cand in cands:
        cand["proxy"] = proxy_relevance(cand["text"], ref)

# ---------------------------------------------------------------------------
# GRID SEARCH
# ---------------------------------------------------------------------------
print("[6/6] Grid search bobot...", flush=True)

# Generate semua kombinasi bobot yang sum <= 1.0, step WEIGHT_STEP
# Normalkan agar w_v + w_b + w_t = 1.0
weight_values = [round(v * WEIGHT_STEP, 2) for v in range(1, int(1.0 / WEIGHT_STEP) + 1)]

combos = []
for wv, wb in product(weight_values, repeat=2):
    wt = round(1.0 - wv - wb, 2)
    if wt < 0.0 or wt > 1.0:
        continue
    if wv + wb + wt < 0.99:   # floating point toleransi
        continue
    combos.append((wv, wb, wt))

print(f"  Total kombinasi: {len(combos)}", flush=True)


def evaluate_weights(w_v, w_b, w_t):
    """Hitung rata-rata Precision@5 (proxy) untuk semua query."""
    p5_list = []
    for cands, ref in zip(all_candidates, references):
        # Hitung final score
        scored = []
        for c in cands:
            fs = w_v * c["vector_score"] + w_b * c["bm25_score"] + w_t * c["taxonomy_score"]
            scored.append((fs, c["proxy"]))
        # Sort desc
        scored.sort(key=lambda x: x[0], reverse=True)
        top5 = scored[:TOP_K]
        # Precision@5: fraksi top-5 yang proxy > threshold
        threshold = 0.1
        p5 = sum(1 for _, prx in top5 if prx >= threshold) / TOP_K
        p5_list.append(p5)
    return sum(p5_list) / len(p5_list)


results = []
for idx, (wv, wb, wt) in enumerate(combos):
    score = evaluate_weights(wv, wb, wt)
    results.append((score, wv, wb, wt))
    if (idx + 1) % 50 == 0:
        print(f"  {idx+1}/{len(combos)} kombinasi dievaluasi...", flush=True)

results.sort(reverse=True)

# Skor bobot saat ini
current_score = evaluate_weights(*CURRENT_W)

# ---------------------------------------------------------------------------
# LAPORAN
# ---------------------------------------------------------------------------
print("\n" + "=" * 65)
print(f"{'Rank':<5} {'w_vector':>9} {'w_bm25':>7} {'w_tax':>6} {'P@5':>8}")
print("-" * 65)

# Tandai bobot saat ini
current_rank = None
for rank, (score, wv, wb, wt) in enumerate(results, 1):
    if (wv, wb, wt) == CURRENT_W:
        current_rank = rank
        break

print(f"[CURRENT]  w_v={CURRENT_W[0]:.2f}  w_b={CURRENT_W[1]:.2f}  w_t={CURRENT_W[2]:.2f}  P@5={current_score:.4f}  (rank {current_rank}/{len(results)})")
print("-" * 65)

for rank, (score, wv, wb, wt) in enumerate(results[:20], 1):
    marker = " ← current" if (wv, wb, wt) == CURRENT_W else ""
    print(f"{rank:<5} {wv:>9.2f} {wb:>7.2f} {wt:>6.2f} {score:>8.4f}{marker}")

# ---------------------------------------------------------------------------
# SIMPAN CSV
# ---------------------------------------------------------------------------
with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(["rank", "w_vector", "w_bm25", "w_taxonomy", "proxy_p5"])
    for rank, (score, wv, wb, wt) in enumerate(results, 1):
        writer.writerow([rank, wv, wb, wt, f"{score:.4f}"])

print(f"\n[SAVED] {OUTPUT_CSV}")
print("[DONE]")
